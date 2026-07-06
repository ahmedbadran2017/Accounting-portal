"""Bank statement import + auto-match reconciliation.

Upload a bank statement (CSV / Excel / PDF), parse it into candidate
transactions, and match each against the account's uncleared book entries
(Payment Entries + Journal Entries) by amount within a date window. The result
splits into:
  • matched       — statement line ↔ a book entry (one click to reconcile)
  • statement_only — on the bank but NOT in the books (missing entry to record)
  • book_only      — in the books but NOT on the statement (still outstanding)

Parsing is format-tolerant: it auto-detects the date / debit / credit / amount
columns from the header, and the caller can override the mapping when a statement
is unusual. PDF is best-effort text extraction. Reconciling the matched lines
reuses the proven reconciliation.mark_bank_cleared (stamps clearance_date, no GL
impact, reversible).
"""
import csv
import io
import re

import frappe
from frappe.utils import flt, getdate

from accounting_portal.api.permissions import assert_portal_access, resolve_companies


def _target(company):
    companies = resolve_companies(company)
    if not companies:
        return None
    return company if (company and company in companies) else companies[0]


# ── number / date parsing ─────────────────────────────────────────────────────

def _parse_amount(v):
    """Money string → float. Handles 1.234,56 and 1,234.56 and (123) negatives."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return flt(v)
    s = str(v).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^\d,.\-]", "", s)
    if not re.search(r"\d", s):
        return None
    # Decide the decimal separator by whichever punctuation comes last.
    last_c, last_d = s.rfind(","), s.rfind(".")
    if last_c > last_d:            # 1.234,56 → comma is decimal
        s = s.replace(".", "").replace(",", ".")
    else:                          # 1,234.56 → dot is decimal
        s = s.replace(",", "")
    try:
        val = float(s)
    except ValueError:
        return None
    return -val if neg else val


def _parse_date(v):
    if v is None:
        return None
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            import datetime
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            continue
    try:
        return str(getdate(s))
    except Exception:
        return None


# ── file extraction ───────────────────────────────────────────────────────────

def _file_bytes(file_url):
    f = frappe.get_doc("File", {"file_url": file_url})
    return f.get_content()


def _rows_from_excel(content):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    return [[c if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]


def _rows_from_csv(content):
    text = content.decode("utf-8", errors="replace") if isinstance(content, bytes) else content
    # sniff delimiter (comma / semicolon / tab)
    sample = text[:4000]
    delim = ";" if sample.count(";") > sample.count(",") else ("\t" if "\t" in sample else ",")
    return [list(r) for r in csv.reader(io.StringIO(text), delimiter=delim)]


def _lines_from_pdf(content):
    text = None
    for mod in ("pdfplumber", "PyPDF2"):
        try:
            if mod == "pdfplumber":
                import pdfplumber
                with pdfplumber.open(io.BytesIO(content)) as pdf:
                    text = "\n".join((p.extract_text() or "") for p in pdf.pages)
            else:
                import PyPDF2
                r = PyPDF2.PdfReader(io.BytesIO(content))
                text = "\n".join((pg.extract_text() or "") for pg in r.pages)
            if text:
                break
        except Exception:
            continue
    if not text:
        frappe.throw("Couldn't read this PDF on the server. Export the statement as Excel/CSV and upload that.")
    # each line → [date, description, amount] guessed by regex
    rows = []
    date_re = re.compile(r"\b(\d{2}[/.\-]\d{2}[/.\-]\d{4}|\d{4}[/.\-]\d{2}[/.\-]\d{2})\b")
    amt_re = re.compile(r"-?\(?\d{1,3}(?:[.,\s]\d{3})*(?:[.,]\d{2})\)?")
    for ln in text.splitlines():
        d = date_re.search(ln)
        if not d:
            continue
        amts = amt_re.findall(ln)
        if not amts:
            continue
        desc = ln[d.end():].strip()
        rows.append([d.group(1), desc, amts[-1]])  # last number = amount/balance heuristic
    return [["Date", "Description", "Amount"]] + rows


_DATE_HINTS = ("date", "تاريخ", "value date", "date valeur", "date opération", "operation date")
_DEBIT_HINTS = ("debit", "débit", "مدين", "withdrawal", "retrait", "sortie", "out")
_CREDIT_HINTS = ("credit", "crédit", "دائن", "deposit", "versement", "entrée", "in")
# "operation"/"opération" is NOT an amount hint — BMCE (and most MA/FR banks) use it
# for the description column; matching it as amount parses narration digits as money.
_AMOUNT_HINTS = ("amount", "montant", "mouvement", "المبلغ", "value")
_DESC_HINTS = ("description", "libelle", "libellé", "label", "narration", "détail", "detail", "wording", "بيان", "reference", "référence", "operation", "opération")


def _detect_columns(header):
    h = [str(x).strip().lower() for x in header]

    def find(hints):
        for i, cell in enumerate(h):
            if any(hint in cell for hint in hints):
                return i
        return None
    cols = {"date": find(_DATE_HINTS), "debit": find(_DEBIT_HINTS),
            "credit": find(_CREDIT_HINTS), "amount": find(_AMOUNT_HINTS),
            "desc": find(_DESC_HINTS)}
    # A statement is either single-signed-amount OR split debit/credit. When the
    # split columns exist they win — a stray "amount" match is usually a text col.
    if cols["debit"] is not None or cols["credit"] is not None:
        cols["amount"] = None
    return cols


@frappe.whitelist()
def parse_statement(file_url=None, mapping=None):
    """Parse an uploaded statement file → candidate transactions + column map.
    `mapping` (optional JSON) overrides the auto-detected column indexes."""
    assert_portal_access()
    if not file_url:
        frappe.throw("No file uploaded")
    content = _file_bytes(file_url)
    ext = file_url.lower().rsplit(".", 1)[-1]
    if ext in ("xlsx", "xlsm"):
        rows = _rows_from_excel(content)
    elif ext == "csv":
        rows = _rows_from_csv(content)
    elif ext == "pdf":
        rows = _lines_from_pdf(content)
    else:
        frappe.throw(f"Unsupported file type .{ext} — use CSV, Excel or PDF.")
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if len(rows) < 2:
        frappe.throw("No rows found in the file.")

    # find the header row = the first row that detects a date column
    header_idx, cols = 0, {}
    for i, r in enumerate(rows[:15]):
        c = _detect_columns(r)
        if c["date"] is not None and (c["amount"] is not None or c["debit"] is not None or c["credit"] is not None):
            header_idx, cols = i, c
            break
    else:
        cols = _detect_columns(rows[0])
    if isinstance(mapping, str):
        import json
        mapping = json.loads(mapping or "{}")
    if mapping:
        # explicit user mapping wins — including explicit nulls ("—" clears a column)
        for k in ("date", "amount", "debit", "credit", "desc"):
            if k in mapping:
                cols[k] = mapping[k]

    def cell(row, idx):
        return row[idx] if (idx is not None and idx < len(row)) else None

    txns = []
    for r in rows[header_idx + 1:]:
        d = _parse_date(cell(r, cols.get("date")))
        if not d:
            continue
        amt = None
        if cols.get("debit") is not None or cols.get("credit") is not None:
            dr = _parse_amount(cell(r, cols.get("debit")))
            cr = _parse_amount(cell(r, cols.get("credit")))
            if dr is not None or cr is not None:
                # credit (money in) positive, debit (money out) negative;
                # banks export debits as positive or already-negative — use magnitude
                amt = (cr or 0) - abs(dr or 0)
        if amt is None:
            amt = _parse_amount(cell(r, cols.get("amount")))
        if amt is None or abs(amt) < 0.005:
            continue
        if abs(amt) > 1e12:  # narration/account-number digits parsed as money
            continue
        txns.append({"date": d, "amount": flt(amt, 2),
                     "description": str(cell(r, cols.get("desc")) or "").strip()[:180]})
    return {"transactions": txns, "count": len(txns), "columns": cols,
            "header": [str(c) for c in rows[header_idx]],
            "preview": [[str(c) for c in row] for row in rows[header_idx:header_idx + 6]]}


# ── matching ──────────────────────────────────────────────────────────────────

@frappe.whitelist()
def match_statement(company=None, account=None, transactions=None, date_window=4):
    """Match parsed statement transactions against the account's uncleared book
    entries by amount (± date_window days). Read-only — proposes what to clear."""
    assert_portal_access()
    target = _target(company)
    if not (target and account):
        frappe.throw("company and account are required")
    if isinstance(transactions, str):
        import json
        transactions = json.loads(transactions or "[]")
    transactions = transactions or []
    window = int(date_window or 4)

    # uncleared book entries on this account (PE + JE), signed +in / −out
    book = frappe.db.sql(
        """SELECT pe.name voucher, 'Payment Entry' doctype, pe.posting_date dt,
                  CASE WHEN pe.paid_to=%(a)s THEN pe.received_amount ELSE -pe.paid_amount END amt,
                  IFNULL(pe.party_name, pe.party) party
           FROM `tabPayment Entry` pe
           WHERE pe.company=%(c)s AND pe.docstatus=1 AND pe.clearance_date IS NULL
             AND (pe.paid_to=%(a)s OR pe.paid_from=%(a)s)
           UNION ALL
           SELECT je.name, 'Journal Entry', je.posting_date,
                  SUM(jea.debit - jea.credit), MAX(IFNULL(jea.party, je.user_remark))
           FROM `tabJournal Entry` je JOIN `tabJournal Entry Account` jea ON jea.parent=je.name
           WHERE je.company=%(c)s AND je.docstatus=1 AND je.clearance_date IS NULL AND jea.account=%(a)s
           GROUP BY je.name""",
        {"c": target, "a": account}, as_dict=True)
    for b in book:
        b["amt"] = flt(b["amt"], 2)
        b["dt"] = str(b["dt"])[:10]
        b["used"] = False

    matched, statement_only = [], []
    for t in transactions:
        tamt = flt(t.get("amount"), 2)
        tdt = getdate(t.get("date"))
        best = None
        for b in book:
            if b["used"] or abs(b["amt"] - tamt) > 0.01:
                continue
            gap = abs((getdate(b["dt"]) - tdt).days)
            if gap <= window and (best is None or gap < best[1]):
                best = (b, gap)
        if best:
            best[0]["used"] = True
            matched.append({"statement": t, "book": {k: best[0][k] for k in ("voucher", "doctype", "dt", "amt", "party")},
                            "day_gap": best[1]})
        else:
            statement_only.append(t)
    book_only = [{k: b[k] for k in ("voucher", "doctype", "dt", "amt", "party")} for b in book if not b["used"]]

    return {
        "company": target, "account": account, "date_window": window,
        "matched": matched, "matched_n": len(matched),
        "statement_only": statement_only, "statement_only_n": len(statement_only),
        "book_only": book_only, "book_only_n": len(book_only),
        "matched_value": flt(sum(m["statement"]["amount"] for m in matched), 2),
        "statement_only_value": flt(sum(t["amount"] for t in statement_only), 2),
        "book_only_value": flt(sum(b["amt"] for b in book_only), 2),
    }
